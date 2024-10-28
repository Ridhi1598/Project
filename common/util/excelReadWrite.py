import openpyxl
class ExcelReadWrite:
    def read_excel_by_column_name(self, fieldName):
        colValues = []
        path = "F:\LCD-Tinna\lcd\OR\Book.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        colNum = None

        rowCount = sheet.max_row
        colCount = sheet.max_column

        for i in range(1, colCount):
            if sheet.cell(row = 1,column = i).value == fieldName:
                colNum = i
                break

        for j in range(2, rowCount):
            colValues.append(sheet.cell(row=j, column = colNum).value)

        return colValues


    def read_excel_by_row_name(self, fieldName):
        rowValues = []
        path = "F:\LCD-Tinna\lcd\OR\Book.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rowNum = None

        rowCount = sheet.max_row
        colCount = sheet.max_column

        for i in range(1, rowCount):
            if str(sheet.cell(row = i,column = 1).value) == fieldName:
                rowNum = i
                break

        for j in range(2, colCount):
            rowValues.append(sheet.cell(row = rowNum, column = j).value)

        return rowValues


    #
    # def excel_write(self , row ,column ,element):
    #     global sheet
    #     global workbook
    #     global path
    #     for i in range(1,row):
    #         for j in range(1,column):
    #            sheet.cell(row=i,column=j).value = element
    #     workbook.save(path)



readExcel = ExcelReadWrite()
print(readExcel.read_excel_by_row_name("3"))